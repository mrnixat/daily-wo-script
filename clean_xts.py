from collections import defaultdict
from pypdf import PdfReader
import shutil
import re 
import os
from datetime import datetime, timedelta
import time
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import vars


SAVE_DIR = os.path.expanduser("~/Downloads/Hapag_XTs")
day_before_folder = os.path.join(SAVE_DIR, "Hapag_XTs_Yesterday")

def clean_xts():
    """
    path = os.path.join(script_dir, "WOSD0001_785752482.pdf")
    work_order = PdfReader(path)

    print (len(work_order.pages)) # prints the number of pages in the PDF
    print(work_order.pages[0].extract_text()) # prints the text of the first page
    print(work_order.pages[1].extract_text()) # prints the text of the second page
    print(work_order.pages[2].extract_text()) # prints the text of the third page
    """

    # CODE TO GO THROUGH FOLDER
    date_pattern = "Date Issue: "
    today = (datetime.now() - timedelta(days=vars.days_back)).strftime("%b %d") # MM/DD/YY
    #           take timedelta out if you want the exact day
    #today = (datetime.now() - timedelta(days=3)).strftime("%b %d")

    date = date_pattern + today
    print(date)

    #move_folder = os.path.join(SAVE_DIR, "Anything_not_of_day")
    update_folder = os.path.join(SAVE_DIR, "Updates")
    #future_folder = os.path.join(SAVE_DIR, "Futures")
    #os.makedirs(move_folder, exist_ok=True)
    os.makedirs(update_folder, exist_ok=True)
    #os.makedirs(future_folder, exist_ok=True)


    """
    # Moves every XT that is not of the day to another folder
    print ("Moving every thats not of the day to another folder...")
    count = 0
    for filename in os.listdir(SAVE_DIR): # for every single XT PDF in the save dir
        if not filename.lower().endswith(".pdf") and not filename.startswith("WOSD"):
            print("Not a PDF/XT, skipping...") # skips if not a PDF and XT
            continue

        path = os.path.join(SAVE_DIR, filename) # saves the path
        work_order = PdfReader(path) # PDFReader Object
        if date not in work_order.pages[0].extract_text(): # if the date we want isn't on the work order...
            try: # try and...
                shutil.move(path, move_folder) # move it to another folder
                print("Moved a PDF that is not of date")
                count+=1
            except Exception as e: # except an error if it doesn't work for some reason
                print("Error")

    print("Moved " + str(count) + " WOs that are not of date to another folder within the SAVE DIR")
    """

    print("waiting...")
    time.sleep(1)

    # This code of block will move every XT that is an update to another folder
    count = 0
    for filename in os.listdir(SAVE_DIR):
        if not filename.lower().endswith(".pdf") and not filename.startswith("WOSD"):
            print("Not a PDF/XT, skipping...") # skips if not a PDF and XT
            continue
        path = os.path.join(SAVE_DIR, filename)
        work_order = PdfReader(path)
        pdf_text = work_order.pages[0].extract_text()
        if "Update" in pdf_text:
            print("Found an update, moving it to another folder...")
            try:
                shutil.move(path, update_folder)
                print("Moved an update")
                count += 1
            except Exception as e:
                print(f"Error moving update: {e}")
    print(f"Moved {count} updates to the update folder")
    print("waiting...")
    time.sleep(1)

    # FOR DAY BEFORE PDFS
    count = 0
    for filename in os.listdir(day_before_folder):
        if not filename.lower().endswith(".pdf") and not filename.startswith("WOSD"):
            print("Not a PDF/XT, skipping...") # skips if not a PDF and XT
            continue
        path = os.path.join(day_before_folder, filename)
        work_order = PdfReader(path)
        pdf_text = work_order.pages[0].extract_text()
        if "Update" in pdf_text:
            print("Found an update, moving it to another folder...")
            try:
                shutil.move(path, update_folder)
                print("Moved an update")
                count += 1
            except Exception as e:
                print(f"Error moving update: {e}")
    print(f"Moved {count} updates to the update folder")
    print("waiting...")
    time.sleep(1)


    count_refs = 0
    count_cons = 0
    WOs = defaultdict(int)
    output = ""
    data = {}

    # Finds the billing refs and container numbers
    for filename in os.listdir(SAVE_DIR): # for every single XT PDF in the save dir
        if not filename.lower().endswith(".pdf") and not filename.startswith("WOSD"):
            print("Not a PDF/XT, skipping...") # skips if not a PDF and XT
            continue

        print("Now on: " + filename)

        path = os.path.join(SAVE_DIR, filename) # saves the path of the file
        if not os.path.exists(path):
            print(f"File {path} does not exist, skipping...")
            continue
        work_order = PdfReader(path) # PDFReader Object
        entire_wo = ""

        for page in work_order.pages: # goes through each page of the PDF
            entire_wo += page.extract_text() # and adds every page of the PDF to a string
        if "TRANSPORT 250.00 USD" not in entire_wo: # checks if the PDF is a Hapag XT or an Import
            print("Possibly not an XT, skipping...\n")
            continue

        text = work_order.pages[0].extract_text() # Extracts the text of the first page
        
    #    if "Update" in text:
    #        print("Found an update")
    #        continue

        billing_pattern = r"Work Order: (\d{9})" # regex for Work Order
        billing_ref = re.findall(billing_pattern, text) # finds the billing ref of the pdf

        if billing_ref[0] in WOs.keys(): # makes sure that every WO is unique
            print("Found duplicate order or an update, skipping this iteration...")
            continue
        output+= filename + "\n"

        WOs[billing_ref[0]]+=1
        count_refs+=1
        output += f"Billing ref: {billing_ref[0]}\n"

        containers = set() # creates a set of containers

        for i in range(len(work_order.pages)): # goes through each and every page of the pdf
            text = work_order.pages[i].extract_text() # extracts the text from each page
            container_pattern = r': \b([A-Z]{4})\s+(\d{6,})\b' # regex for container numbers
            container_nums = re.findall(container_pattern, text) # finds every single container number on that page
            for c in container_nums: # iterates through all of the found container numbers
                container = ''.join(c) # deletes the space
                containers.add(container) # adds them into the set

        print(billing_ref)
        print(containers)
        output += f"Containers: {', '.join(containers)}\n\n"
        print("There is/are " + str(len(containers)) + " container(s) in this WO")
        print()
        count_cons = count_cons + len(containers)
        data[billing_ref[0]] = ', '.join(containers)

    print("Found " + str(count_refs) + " billing refs and found " + str(count_cons) + " containers")
    print(data)

    count_refs = 0
    count_cons = 0
    WOs = defaultdict(int)
    output = ""
    data_before = {}

    # DOES SAME THING FOR THE DAY BEFORE PDFs
    for filename in os.listdir(day_before_folder): # for every single XT PDF in the save dir
        if not filename.lower().endswith(".pdf") and not filename.startswith("WOSD"):
            print("Not a PDF/XT, skipping...") # skips if not a PDF and XT
            continue

        print("Now on: " + filename)

        path = os.path.join(SAVE_DIR, day_before_folder, filename) # saves the path of the file
        if not os.path.exists(path):
            print(f"File {path} does not exist, skipping...")
            continue
        work_order = PdfReader(path) # PDFReader Object
        entire_wo = ""

        for page in work_order.pages: # goes through each page of the PDF
            entire_wo += page.extract_text() # and adds every page of the PDF to a string
        if "TRANSPORT 250.00 USD" not in entire_wo: # checks if the PDF is a Hapag XT or an Import
            print("Possibly not an XT, skipping...\n")
            continue

        text = work_order.pages[0].extract_text() # Extracts the text of the first page
        
    #    if "Update" in text:
    #        print("Found an update")
    #        continue

        billing_pattern = r"Work Order: (\d{9})" # regex for Work Order
        billing_ref = re.findall(billing_pattern, text) # finds the billing ref of the pdf

        if billing_ref[0] in WOs.keys(): # makes sure that every WO is unique
            print("Found duplicate order or an update, skipping this iteration...")
            continue
        output+= filename + "\n"

        WOs[billing_ref[0]]+=1
        count_refs+=1
        output += f"Billing ref: {billing_ref[0]}\n"

        containers = set() # creates a set of containers

        for i in range(len(work_order.pages)): # goes through each and every page of the pdf
            text = work_order.pages[i].extract_text() # extracts the text from each page
            container_pattern = r': \b([A-Z]{4})\s+(\d{6,})\b' # regex for container numbers
            container_nums = re.findall(container_pattern, text) # finds every single container number on that page
            for c in container_nums: # iterates through all of the found container numbers
                container = ''.join(c) # deletes the space
                containers.add(container) # adds them into the set

        print(billing_ref)
        print(containers)
        output += f"Containers: {', '.join(containers)}\n\n"
        print("There is/are " + str(len(containers)) + " container(s) in this WO")
        print()
        count_cons = count_cons + len(containers)
        data_before[billing_ref[0]] = ', '.join(containers)

    print("Found " + str(count_refs) + " billing refs and found " + str(count_cons) + " containers from day before.")
    print(data_before)









    """
    with open(os.path.join(SAVE_DIR, "output.csv"), "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Billing Ref", "Containers"])

        for billing_ref, containers in data.items():
            string_ref = str(billing_ref)
            writer.writerow([string_ref, containers])
    """


    multi_fills = [
        PatternFill(start_color="e0e4f4", end_color="e0e4f4", fill_type="solid"),
        PatternFill(start_color="ffecdc", end_color="ffecdc", fill_type="solid")   
    ]

    date_today = (datetime.now() - timedelta(days=vars.days_back)).strftime("%m-%d-%Y")
    wb = Workbook()
    ws = wb.active
    ws.title = "Hapag XTs " + date_today

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)

    # Write header
    ws.append(["Billing Ref", "Containers"])

    fill_index = 0  # Start with first multi-fill

    for col in range(1, 3):  # Assuming 2 columns
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    for billing_ref, container_string in data.items():
        containers = container_string.split(', ')
        is_multi = len(containers) > 1

        fill = None
        if is_multi:
            fill = multi_fills[fill_index % len(multi_fills)]
            fill_index += 1  # Alternate only when there's a new multi

        for container in containers:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=billing_ref)
            ws.cell(row=row, column=2, value=container)
            if fill:
                ws.cell(row=row, column=1).fill = fill
                ws.cell(row=row, column=2).fill = fill

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 3


    # Create new worksheet for Day Before
    ws_day_before = wb.create_sheet(title="XTs Day before past 4PM")
    ws_day_before.append(["Billing Ref", "Containers"])

    # Format header for ws_day_before
    for col in range(1, 3):
        cell = ws_day_before.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    fill_index = 0
    for billing_ref, container_string in data_before.items():
        containers = container_string.split(', ')
        is_multi = len(containers) > 1

        fill = None
        if is_multi:
            fill = multi_fills[fill_index % len(multi_fills)]
            fill_index += 1

        for container in containers:
            row = ws_day_before.max_row + 1
            ws_day_before.cell(row=row, column=1, value=billing_ref)
            ws_day_before.cell(row=row, column=2, value=container)
            if fill:
                ws_day_before.cell(row=row, column=1).fill = fill
                ws_day_before.cell(row=row, column=2).fill = fill

    # Auto-fit columns for ws_day_before
    for col in ws_day_before.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_day_before.column_dimensions[col_letter].width = max_len + 3

    filename = f"Hapag XTs {date_today}.xlsx" 
    wb.save(os.path.join(SAVE_DIR, filename)) # saving the workbook
    print(f"Output written to {filename} in the save directory")